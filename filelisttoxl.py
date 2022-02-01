import openpyxl
import os
import sys
import glob

def getFileInfo(filepath, option=1):
    if option == 1: # only filename
        return os.path.split(filepath)[1]
    else: # fullpath
        return filepath



def createFileList():
    xlrow = 1
    xlcolumn = 1

    # create book file
    book = openpyxl.Workbook()

    # get sheet
    sheet = book['Sheet']
    
    # target path
    path = sys.argv[1]

    sheet.cell(xlrow, xlcolumn).value = "Directory:" + path
    xlrow += 1

    # get file list
    files = glob.glob(path + "/*")
    for file in files:
        sheet.cell(xlrow, xlcolumn).value = getFileInfo(file)
        xlrow += 1

    # rename sheet
    sheet.title = 'filelist'

    # save file
    book.save("filelist.xlsx")

def dispDesc():
    print('-- filelistxl --')
    print('>python3 filelistxl [path]')
   
# main flow
def main():
    argc = len(sys.argv)
    if argc == 1:
        dispDesc()
        sys.exit(0)
    
    if os.path.exists(sys.argv[1]) == False:
        dispDesc()
        sys.exit(0)

    createFileList()



if __name__ == "__main__":
    main()

